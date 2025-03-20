import pandas as pd
import yaml
import json
import csv
import jsonschema
from jsonschema import validate, ValidationError, SchemaError
from io import StringIO
from datetime import datetime
from pathlib import Path
from tabulate import tabulate
from utils.file_naming import FileNamingConvention
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import logging

class LectorArchivos:
    FORMATOS_SOPORTADOS = {
        '.csv': pd.read_csv,
        '.xls': pd.read_excel,
        '.xlsx': pd.read_excel,
        '.tsv': lambda f: pd.read_csv(f, sep='\t'),
        '.ods': pd.read_excel,
        '.json': pd.read_json,
        '.yml': lambda f: pd.DataFrame(yaml.safe_load(open(f))),
        '.yaml': lambda f: pd.DataFrame(yaml.safe_load(open(f))),
        '.html': lambda f: pd.read_html(f)[0]
    }

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)

    def listar_archivos_input(self, clinic_path, tipo_archivo=None):
        """
        Lista los archivos disponibles en la carpeta input de la clínica actual
        
        Args:
            clinic_path: Ruta base de la clínica
            tipo_archivo: Tipo de archivo a buscar (pacientes, FARC, BIO, MTP)
        """
        input_path = clinic_path / "lector_archivos" / "input"

        if not input_path.exists():
            print(f"\nNo se encontró la carpeta de entrada: {input_path}")
            return None

        # Buscar archivos con las extensiones soportadas
        archivos = []
        for ext in self.FORMATOS_SOPORTADOS.keys():
            if tipo_archivo:
                # Si se especifica tipo, buscar archivos que contengan ese tipo en el nombre
                pattern = f"*{tipo_archivo}*{ext}"
            else:
                pattern = f"*{ext}"
            archivos.extend(input_path.glob(pattern))

        if not archivos:
            tipo_msg = f" de tipo {tipo_archivo}" if tipo_archivo else ""
            print(f"\nNo se encontraron archivos{tipo_msg} en: {input_path}")
            return None

        print(f"\n=== ARCHIVOS DISPONIBLES EN {input_path} ===")
        for idx, archivo in enumerate(archivos, 1):
            print(f"{idx}. {archivo.name}")

        while True:
            try:
                opcion = int(input("\nSeleccione el archivo a procesar (0 para cancelar): ")) - 1
                if opcion == -1:
                    return None
                if 0 <= opcion < len(archivos):
                    return archivos[opcion]
                print("Opción no válida")
            except ValueError:
                print("Por favor ingrese un número válido")

    def leer_archivo(self, file_path):
        """Lee un archivo y retorna un DataFrame junto con su extensión"""
        file_path = Path(file_path)
        extension = file_path.suffix.lower()
        if extension == '.xml':
            self.logger.info(f"Procesando archivo XML: {file_path}")
            df = self.leer_xml(file_path)
            return df, extension
        if extension not in self.FORMATOS_SOPORTADOS:
            raise ValueError(f"Formato no soportado: {extension}")
        
        try:
            df = self.FORMATOS_SOPORTADOS[extension](file_path)
            return df, extension
        except Exception as e:
            raise Exception(f"Error al leer archivo: {str(e)}")

    def leer_xml(self, ruta_archivo):
        import xml.etree.ElementTree as ET
        try:
            tree = ET.parse(ruta_archivo)
            root = tree.getroot()
            data = []
            for child in root:
                record = {}
                for elem in child:
                    record[elem.tag] = elem.text
                data.append(record)
            df = pd.DataFrame(data)
            self.logger.info(f"Archivo XML leído correctamente: {ruta_archivo}")
            return df
        except Exception as e:
            self.logger.error(f"Error al leer XML: {e}")
            raise e

    def leer_archivo_mixto(self, ruta_archivo):
        self.logger.info(f"Procesando archivo mixto: {ruta_archivo}")
        try:
            df = self.leer_json(ruta_archivo)
            self.logger.info("Archivo procesado como JSON.")
            return df
        except Exception:
            try:
                df = self.leer_xml(ruta_archivo)
                self.logger.info("Archivo procesado como XML.")
                return df
            except Exception:
                try:
                    df = pd.read_csv(ruta_archivo)
                    self.logger.info("Archivo procesado como CSV.")
                    return df
                except Exception as e:
                    self.logger.error(f"Error al procesar archivo mixto: {e}")
                    raise e

    def procesar_lote(self, lista_archivos, output_dir, clinic_initials=None, pre_hooks=None, post_hooks=None):
        resultados = []
        for ruta in lista_archivos:
            try:
                df, estructura = self.procesar_archivo(ruta, output_dir, clinic_initials, pre_hooks, post_hooks)
                resultados.append((ruta, df, estructura))
                self.logger.info(f"Procesado correctamente: {ruta}")
            except Exception as e:
                self.logger.error(f"Error al procesar {ruta}: {e}")
        return resultados

    def validar_datos_genericos(self, df, reglas):
        # reglas es un dict: clave => función de validación (retorna True si es válido)
        errores = {}
        for col, func in reglas.items():
            if col in df.columns:
                invalidos = df[~df[col].apply(func)]
                if not invalidos.empty:
                    errores[col] = invalidos.index.tolist()
                    self.logger.warning(f"Datos inválidos en columna {col}: {errores[col]}")
        return errores

    def analizar_estructura(self, df):
        """Analiza y retorna la estructura detallada del DataFrame"""
        estructura = []
        for columna, tipo in df.dtypes.items():
            # Mantener el valor exacto incluyendo nulos
            ejemplo = df[columna].iloc[0]
            # No convertir a string para mantener el tipo original
            info_campo = {
                'nombre': columna,
                'tipo': str(tipo),
                'ejemplo': ejemplo,
                'n_unicos': df[columna].nunique(),
                'tiene_nulos': df[columna].isnull().any(),
                'tipo_nulos': 'NaN' if df[columna].dtype.kind in 'fc' else 
                            'NULL' if df[columna].dtype.kind == 'O' else 
                            'Vacío'
            }
            
            if info_campo['n_unicos'] <= 5:
                # Mantener los valores exactos sin convertir
                info_campo['valores_unicos'] = df[columna].unique().tolist()
            
            estructura.append(info_campo)
        
        return estructura

    def mostrar_estructura(self, estructura):
        """Muestra la estructura en formato tabular"""
        print("\n=== ESTRUCTURA DEL ARCHIVO ===")
        print("=" * 120)
        
        formato = "{:<35} | {:<12} | {:<35} | {:<25}"
        print(formato.format("NOMBRE DEL CAMPO", "TIPO", "EJEMPLO", "INFO ADICIONAL"))
        print("-" * 120)
        
        for campo in estructura:
            ejemplo = str(campo['ejemplo'])[:35] if pd.notna(campo['ejemplo']) else '<NULO>'
            info_adicional = f"Nulos: {campo['tipo_nulos']}" if campo['tiene_nulos'] else "No tiene nulos"
            
            print(formato.format(
                campo['nombre'][:35],
                campo['tipo'],
                ejemplo,
                info_adicional
            ))
            
            if 'valores_unicos' in campo:
                print(f"   • Valores posibles ({campo['n_unicos']}): ")
                for valor in campo['valores_unicos']:
                    print(f"     - {str(valor) if pd.notna(valor) else '<NULO>'}")
            print("-" * 120)

    def modificar_estructura(self, estructura):
        """Permite al usuario modificar la estructura interactivamente"""
        while True:
            self.mostrar_estructura(estructura)
            if input("\n¿Desea modificar algún campo? (S/N): ").upper() != 'S':
                break
                
            print("\nCampos disponibles:")
            for idx, campo in enumerate(estructura, 1):
                print(f"{idx}. {campo['nombre']}")
            
            try:
                num_campo = int(input("\nSeleccione el número del campo a modificar (0 para terminar): ")) - 1
                if num_campo == -1:
                    break
                if 0 <= num_campo < len(estructura):
                    campo = estructura[num_campo]
                    nuevo_nombre = input(f"Nuevo nombre para '{campo['nombre']}' (Enter para mantener): ")
                    if nuevo_nombre:
                        campo['nombre'] = nuevo_nombre
                else:
                    print("Número de campo no válido")
            except ValueError:
                print("Por favor ingrese un número válido")
        
        return estructura

    def _limpiar_valor_para_json(self, valor):
        """Limpia un valor para que sea compatible con JSON"""
        if pd.isna(valor):
            return None
        if isinstance(valor, (list, tuple)):
            return [self._limpiar_valor_para_json(v) for v in valor]
        return valor

    def preguntar_formato_salida(self):
        """Permite al usuario seleccionar el formato de salida"""
        FORMATOS_SALIDA = {
            '1': ('CSV', lambda df, p: df.to_csv(p, index=False), '.csv'),
            '2': ('Excel', lambda df, p: df.to_excel(p, index=False), '.xlsx'),
            '3': ('JSON', lambda df, p: df.to_json(p, orient='records', indent=2), '.json'),
            '4': ('HTML', lambda df, p: df.to_html(p, index=False), '.html'),
            '5': ('YAML', lambda df, p: yaml.dump(yaml.safe_load(df.to_json(orient='records')), open(p, 'w')), '.yaml'),
            '6': ('TSV', lambda df, p: df.to_csv(p, sep='\t', index=False), '.tsv'),
            '7': ('ODS', lambda df, p: df.to_excel(p, engine='odf', index=False), '.ods')
        }

        print("\n=== FORMATOS DE SALIDA DISPONIBLES ===")
        for key, (nombre, _, _) in FORMATOS_SALIDA.items():
            print(f"{key}. {nombre}")
        
        while True:
            opcion = input("\nSeleccione el formato de salida (1-7): ")
            if opcion in FORMATOS_SALIDA:
                return FORMATOS_SALIDA[opcion]
            print("Opción no válida")

    def generar_documentacion(self, estructura, clinic_initials, tipo_archivo, output_dir):
        """Genera archivo de documentación con la estructura validada"""
        df_estructura = pd.DataFrame([{
            'Campo': campo['nombre'],
            'Tipo': campo['tipo'],
            'Ejemplo': campo['ejemplo'],
            'Valores_Unicos': campo['n_unicos'],
            'Valores_Posibles': str(campo.get('valores_unicos', []))
        } for campo in estructura])

        # Solicitar formato de salida
        nombre_formato, exportador, extension = self.preguntar_formato_salida()
        
        # Usar la convención estándar para nombrar el archivo
        filename = FileNamingConvention.generate_filename(
            clinic_initials=clinic_initials,
            module=tipo_archivo,
            extension=extension[1:]  # Remover el punto de la extensión
        )
        
        output_path = FileNamingConvention.get_export_path(output_dir, filename)

        try:
            exportador(df_estructura, output_path)
            print(f"\nDocumentación guardada en: {output_path}")
            
            if input("\n¿Desea exportar en otro formato? (S/N): ").upper() == 'S':
                self.generar_documentacion(estructura, clinic_initials, tipo_archivo, output_dir)
            
            return output_path
        except Exception as e:
            print(f"Error al exportar documentación: {str(e)}")
            return None

    def procesar_archivo(self, file_path, output_dir, clinic_initials=None, pre_hooks=None, post_hooks=None):
        """Procesa un archivo: lee, analiza, valida y documenta su estructura con pipeline configurable"""
        clinic_path = Path(output_dir).parent.parent

        # Ejecutar hooks previos si están definidos
        if pre_hooks:
            for hook in pre_hooks:
                hook(file_path)  # Se espera que hook sea una función que reciba file_path
        
        # Si no se proporcionó un archivo específico, mostrar lista de archivos disponibles
        if not isinstance(file_path, Path):
            tipo_archivo = None
            if clinic_initials:
                # Intentar determinar el tipo de archivo basado en la extensión
                tipo_archivo = Path(file_path).suffix.lower()
                if tipo_archivo not in self.FORMATOS_SOPORTADOS:
                    tipo_archivo = None
            archivo_seleccionado = self.listar_archivos_input(clinic_path, tipo_archivo)
            if archivo_seleccionado is None:
                return None, None
            file_path = archivo_seleccionado

        print(f"\nProcesando archivo: {file_path}")

        try:
            # Leer archivo y mejorar la detección del tipo mediante la extensión
            df, extension = self.leer_archivo(file_path)
            print(f"Archivo leído correctamente con formato {extension}")

            # Analizar estructura
            estructura = self.analizar_estructura(df)
            print("\nEstructura inicial detectada:")
            self.mostrar_estructura(estructura)

            # Permitir modificaciones
            estructura = self.modificar_estructura(estructura)

            # Confirmar estructura final
            print("\nEstructura final:")
            self.mostrar_estructura(estructura)
            
            if input("\n¿La estructura final es correcta? (S/N): ").upper() == 'S':
                # Ejecutar hooks posteriores si están definidos
                if post_hooks:
                    for hook in post_hooks:
                        hook(file_path, estructura)   # Se espera que hook reciba file_path y la estructura

                # Generar documentación
                if clinic_initials:
                    tipo_archivo = Path(file_path).stem.split('_')[0]
                    self.generar_documentacion(estructura, clinic_initials, tipo_archivo, output_dir)
                return df, estructura
            else:
                print("\nProceso cancelado por el usuario")
                return None, None

        except Exception as e:
            raise Exception(f"Error al procesar archivo: {str(e)}")
    
    def leer_csv(self, ruta_archivo, encoding='utf-8', delimiter=None, 
                 headers='auto', encoding_fallbacks=None, **kwargs):
        """
        Lee un archivo CSV con soporte para múltiples codificaciones, 
        detección automática de delimitadores y manejo de encabezados personalizados.
        
        Args:
            ruta_archivo (str): Ruta al archivo CSV a leer
            encoding (str, optional): Codificación del archivo. Por defecto 'utf-8'
            delimiter (str, optional): Delimitador a usar. Si es None, se detecta automáticamente
            headers (str/list, optional): 
                - 'auto': usa primera fila como encabezados (predeterminado)
                - 'none': no usa encabezados 
                - list: lista de strings con encabezados personalizados
            encoding_fallbacks (list, optional): Lista de codificaciones alternativas a intentar
                                               si la principal falla
            **kwargs: Argumentos adicionales para pandas.read_csv
            
        Returns:
            pandas.DataFrame: Datos del CSV convertidos a DataFrame
            
        Raises:
            ValueError: Si no se puede leer el archivo con ninguna de las codificaciones
        """
        # Lista de codificaciones a intentar
        if encoding_fallbacks is None:
            encoding_fallbacks = ['latin-1', 'iso-8859-1', 'cp1252']
        
        # Asegurar que la codificación principal está al inicio
        encodings_to_try = [encoding]
        for enc in encoding_fallbacks:
            if enc != encoding:
                encodings_to_try.append(enc)
        
        # Intentar diferentes codificaciones
        last_error = None
        for enc in encodings_to_try:
            try:
                # Si el delimitador no está especificado, detectarlo automáticamente
                if delimiter is None:
                    delimiter = self._detectar_delimitador(ruta_archivo, enc)
                
                # Configurar manejo de encabezados
                if headers == 'auto':
                    header = 0  # Primera fila como encabezados
                elif headers == 'none':
                    header = None  # Sin encabezados
                elif isinstance(headers, list):
                    # Usar nombres de columnas personalizados
                    kwargs['names'] = headers
                    # Si se proporcionan nombres personalizados, saltamos la primera fila
                    # solo si no hemos especificado explícitamente header en kwargs
                    if 'header' not in kwargs:
                        header = 0
                    else:
                        header = kwargs.pop('header')
                else:
                    raise ValueError("El parámetro 'headers' debe ser 'auto', 'none' o una lista")
                
                # Aplicar header si no está en kwargs
                if 'header' not in kwargs:
                    kwargs['header'] = header
                
                # Leer el CSV con pandas
                return pd.read_csv(ruta_archivo, encoding=enc, sep=delimiter, **kwargs)
            
            except Exception as e:
                last_error = e
                continue
        
        # Si llegamos aquí, ninguna codificación funcionó
        raise ValueError(f"No se pudo leer el archivo CSV con ninguna codificación. "
                         f"Último error: {str(last_error)}")
    
    def _detectar_delimitador(self, ruta_archivo, encoding='utf-8'):
        """
        Detecta automáticamente el delimitador de un archivo CSV.
        
        Args:
            ruta_archivo (str): Ruta al archivo CSV
            encoding (str): Codificación del archivo
            
        Returns:
            str: Delimitador detectado, ',' por defecto si no se puede detectar
        """
        try:
            with open(ruta_archivo, 'r', encoding=encoding) as archivo:
                # Leer las primeras 5 líneas para la detección
                sample = ''.join(archivo.readline() for _ in range(5))
                
                if sample.strip():
                    dialect = csv.Sniffer().sniff(sample)
                    return dialect.delimiter
                return ','  # Valor predeterminado si no se puede detectar
        except Exception:
            return ','  # Valor predeterminado en caso de error

    def leer_excel(self, ruta_archivo, sheet_name=0, range_spec=None, 
                   calc_formulas=True, parse_formats=True, ignore_hidden=False, 
                   **kwargs):
        """
        Lee un archivo Excel con soporte avanzado para selección de hojas,
        rangos específicos de datos y manejo de formatos/fórmulas.
        
        Args:
            ruta_archivo (str): Ruta al archivo Excel a leer
            sheet_name (str/int/list, optional): 
                - str: Nombre de la hoja
                - int: Índice de la hoja (0-based)
                - list: Lista de nombres o índices para leer múltiples hojas
                - None: Lee todas las hojas
            range_spec (str/dict, optional): Especificación del rango de datos a leer
                - str: Rango en notación de Excel (ej. "A1:C10")
                - dict: Con claves 'start_row', 'end_row', 'start_col', 'end_col'
            calc_formulas (bool, optional): Si es True, evalúa las fórmulas en las celdas
            parse_formats (bool, optional): Si es True, interpreta los formatos de celda
            ignore_hidden (bool, optional): Si es True, ignora filas/columnas ocultas
            **kwargs: Argumentos adicionales para pandas.read_excel
        
        Returns:
            pandas.DataFrame o dict: Si sheet_name es un solo valor, retorna un DataFrame.
                                    Si es una lista o None, retorna un dict de DataFrames.
        
        Raises:
            ValueError: Si el archivo no existe o hay un error al leer
            KeyError: Si la hoja especificada no existe en el archivo
        """
        try:
            # Manejar el rango si está en formato Excel (A1:C10)
            if isinstance(range_spec, str):
                # Convertir notación de Excel a índices de filas y columnas
                range_dict = self._convertir_rango_excel(range_spec)
                
                if 'skiprows' not in kwargs:
                    kwargs['skiprows'] = range_dict['start_row']
                
                if 'nrows' not in kwargs:
                    # nrows es la cantidad de filas a leer, no el índice final
                    kwargs['nrows'] = range_dict['end_row'] - range_dict['start_row'] + 1
                
                # Para columnas, usamos usecols
                if 'usecols' not in kwargs:
                    # Convertir índices numéricos de columna a letras de Excel
                    start_col_letter = get_column_letter(range_dict['start_col'] + 1)  # +1 porque es 1-based
                    end_col_letter = get_column_letter(range_dict['end_col'] + 1)
                    
                    # Crear una función que filtre columnas en el rango especificado
                    def filtro_columnas(col):
                        if isinstance(col, int):
                            return range_dict['start_col'] <= col <= range_dict['end_col']
                        elif isinstance(col, str):
                            col_idx = column_index_from_string(col) - 1  # -1 porque es 0-based
                            return range_dict['start_col'] <= col_idx <= range_dict['end_col']
                        return False
                    
                    kwargs['usecols'] = filtro_columnas
            
            elif isinstance(range_spec, dict):
                # Si ya tenemos un diccionario con los índices, usarlo directamente
                if 'start_row' in range_spec and 'skiprows' not in kwargs:
                    kwargs['skiprows'] = range_spec['start_row']
                
                if 'end_row' in range_spec and 'start_row' in range_spec and 'nrows' not in kwargs:
                    kwargs['nrows'] = range_spec['end_row'] - range_spec['start_row'] + 1
                
                if 'start_col' in range_spec and 'end_col' in range_spec and 'usecols' not in kwargs:
                    start_col = range_spec['start_col']
                    end_col = range_spec['end_col']
                    
                    # Crear una función que filtre columnas en el rango especificado
                    def filtro_columnas(col):
                        if isinstance(col, int):
                            return start_col <= col <= end_col
                        elif isinstance(col, str):
                            col_idx = column_index_from_string(col) - 1
                            return start_col <= col_idx <= end_col
                        return False
                    
                    kwargs['usecols'] = filtro_columnas
            
            # Configurar opciones para manejo de formatos y fórmulas
            if 'engine' not in kwargs:
                kwargs['engine'] = 'openpyxl'  # openpyxl tiene mejor soporte para formatos
            
            converters = kwargs.get('converters', {})
            
            # Si se solicita evaluar fórmulas
            if calc_formulas:
                kwargs['engine_kwargs'] = kwargs.get('engine_kwargs', {})
                kwargs['engine_kwargs']['data_only'] = True
            
            # Manejar formatos de celda específicos si se ha solicitado
            if parse_formats:
                # Este pre-procesamiento nos permite detectar y manejar formatos específicos
                formatos = self._obtener_formatos_excel(ruta_archivo, sheet_name)
                # Añadimos los converters basados en los formatos detectados
                converters.update(self._crear_converters_por_formato(formatos))
                kwargs['converters'] = converters
            
            # Manejar elementos ocultos
            if ignore_hidden:
                filas_ocultas, cols_ocultas = self._detectar_elementos_ocultos(ruta_archivo, sheet_name)
                
                # Ignorar filas ocultas
                if filas_ocultas and 'skiprows' not in kwargs:
                    original_skiprows = kwargs.get('skiprows', None)
                    if original_skiprows is None:
                        kwargs['skiprows'] = filas_ocultas
                    elif callable(original_skiprows):
                        # Si skiprows es una función, la envolvemos
                        original_func = original_skiprows
                        kwargs['skiprows'] = lambda x: x in filas_ocultas or original_func(x)
                    elif isinstance(original_skiprows, int):
                        # Si es un entero, convertirlo a lista
                        kwargs['skiprows'] = list(range(original_skiprows)) + filas_ocultas
                    elif isinstance(original_skiprows, list):
                        kwargs['skiprows'] = original_skiprows + filas_ocultas
                
                # Ignorar columnas ocultas
                if cols_ocultas and 'usecols' not in kwargs:
                    original_usecols = kwargs.get('usecols', None)
                    if original_usecols is None:
                        # Crear una función que filtre columnas no ocultas
                        kwargs['usecols'] = lambda col: col not in cols_ocultas
                    elif callable(original_usecols):
                        # Si usecols es una función, la envolvemos
                        original_func = original_usecols
                        kwargs['usecols'] = lambda col: col not in cols_ocultas and original_func(col)
                    elif isinstance(original_usecols, list):
                        # Si es una lista, filtrar las columnas ocultas
                        kwargs['usecols'] = [col for col in original_usecols if col not in cols_ocultas]
            
            # Leer el archivo con pandas
            return pd.read_excel(ruta_archivo, sheet_name=sheet_name, **kwargs)
        
        except Exception as e:
            raise ValueError(f"Error al leer archivo Excel: {str(e)}")
    
    def _convertir_rango_excel(self, range_str):
        """
        Convierte un rango en formato Excel (A1:C10) a índices numéricos.
        
        Args:
            range_str (str): Rango en notación de Excel
        
        Returns:
            dict: Diccionario con índices start_row, end_row, start_col, end_col
        """
        # Validar formato del rango (A1:C10)
        if not re.match(r'^[A-Za-z]+\d+:[A-Za-z]+\d+$', range_str):
            raise ValueError(f"Formato de rango inválido: {range_str}, debe ser como 'A1:C10'")
        
        # Separar las celdas de inicio y fin
        start_cell, end_cell = range_str.split(':')
        
        # Separar letras y números de cada celda
        start_col_str = ''.join(c for c in start_cell if c.isalpha())
        start_row_str = ''.join(c for c in start_cell if c.isdigit())
        end_col_str = ''.join(c for c in end_cell if c.isalpha())
        end_row_str = ''.join(c for c in end_cell if c.isdigit())
        
        # Convertir letras de columna a índices (0-based)
        start_col = column_index_from_string(start_col_str) - 1
        end_col = column_index_from_string(end_col_str) - 1
        
        # Convertir filas a índices (0-based)
        start_row = int(start_row_str) - 1
        end_row = int(end_row_str) - 1
        
        return {
            'start_row': start_row,
            'end_row': end_row,
            'start_col': start_col,
            'end_col': end_col
        }
    
    def _obtener_formatos_excel(self, ruta_archivo, sheet_name):
        """
        Obtiene los formatos de celda de una hoja Excel.
        
        Args:
            ruta_archivo (str): Ruta del archivo Excel
            sheet_name (str/int): Nombre o índice de la hoja
        
        Returns:
            dict: Diccionario con información de formatos por columna
        """
        formatos = {}
        try:
            # Abrir el libro con openpyxl para acceder a los formatos
            workbook = openpyxl.load_workbook(ruta_archivo, data_only=False)
            
            # Obtener la hoja correcta
            if isinstance(sheet_name, int):
                # Si es un índice
                sheet = workbook.worksheets[sheet_name]
            elif isinstance(sheet_name, str):
                # Si es un nombre
                sheet = workbook[sheet_name]
            else:
                # Si es None o una lista, usar la primera hoja
                sheet = workbook.active
            
            # Analizar la primera fila con datos para detectar formatos
            # (asumiendo que la primera fila puede ser encabezados)
            for row in sheet.iter_rows(min_row=2, max_row=10):  # Examinar algunas filas
                for cell in row:
                    col_idx = cell.column - 1  # Convertir a 0-based
                    
                    # Inicializar la entrada para esta columna si no existe
                    if col_idx not in formatos:
                        formatos[col_idx] = {'tipo': None, 'formato': None}
                    
                    # Detectar tipo de dato y formato
                    if cell.data_type == 'n':  # Número
                        if cell.number_format and ('/' in cell.number_format or 
                                                  'y' in cell.number_format.lower() or 
                                                  'm' in cell.number_format.lower() or 
                                                  'd' in cell.number_format.lower()):
                            formatos[col_idx] = {'tipo': 'fecha', 'formato': cell.number_format}
                        elif cell.number_format and '%' in cell.number_format:
                            formatos[col_idx] = {'tipo': 'porcentaje', 'formato': cell.number_format}
                        elif cell.number_format and ('$' in cell.number_format or 
                                                   '€' in cell.number_format or 
                                                   '£' in cell.number_format):
                            formatos[col_idx] = {'tipo': 'moneda', 'formato': cell.number_format}
                        else:
                            formatos[col_idx] = {'tipo': 'numero', 'formato': cell.number_format}
                    elif cell.data_type == 'd':  # Fecha
                        formatos[col_idx] = {'tipo': 'fecha', 'formato': cell.number_format}
                    elif cell.data_type == 'f':  # Fórmula
                        formatos[col_idx] = {'tipo': 'formula', 'formato': None}
                    elif cell.data_type == 'b':  # Booleano
                        formatos[col_idx] = {'tipo': 'booleano', 'formato': None}
            
            return formatos
            
        except Exception as e:
            # Si hay error, retornar un diccionario vacío
            return {}
    
    def _crear_converters_por_formato(self, formatos):
        """
        Crea funciones de conversión basadas en los formatos detectados.
        
        Args:
            formatos (dict): Diccionario con información de formatos
        
        Returns:
            dict: Diccionario de funciones de conversión para cada columna
        """
        converters = {}
        
        for col_idx, info in formatos.items():
            if info['tipo'] == 'fecha':
                # Función para convertir a datetime
                converters[col_idx] = lambda x: pd.to_datetime(x, errors='coerce')
            elif info['tipo'] == 'porcentaje':
                # Función para manejar porcentajes
                converters[col_idx] = lambda x: float(x) / 100 if pd.notnull(x) else x
            elif info['tipo'] == 'moneda':
                # Función para limpiar símbolos de moneda
                def clean_currency(val):
                    if pd.isna(val):
                        return val
                    if isinstance(val, (int, float)):
                        return val
                    return float(''.join(c for c in str(val) if c.isdigit() or c in '.-'))
                converters[col_idx] = clean_currency
        
        return converters
    
    def _detectar_elementos_ocultos(self, ruta_archivo, sheet_name):
        """
        Detecta filas y columnas ocultas en una hoja Excel.
        
        Args:
            ruta_archivo (str): Ruta del archivo Excel
            sheet_name (str/int): Nombre o índice de la hoja
        
        Returns:
            tuple: (filas_ocultas, columnas_ocultas)
        """
        filas_ocultas = []
        columnas_ocultas = []
        
        try:
            workbook = openpyxl.load_workbook(ruta_archivo)
            
            # Obtener la hoja correcta
            if isinstance(sheet_name, int):
                sheet = workbook.worksheets[sheet_name]
            elif isinstance(sheet_name, str):
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            
            # Detectar filas ocultas
            for i, row in enumerate(sheet.rows):
                if sheet.row_dimensions[i + 1].hidden:
                    filas_ocultas.append(i)
            
            # Detectar columnas ocultas
            for i in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(i)
                if sheet.column_dimensions[col_letter].hidden:
                    columnas_ocultas.append(i - 1)  # Convertir a 0-based
                    columnas_ocultas.append(col_letter)
            
            return filas_ocultas, columnas_ocultas
            
        except Exception:
            # Si hay error, retornar listas vacías
            return [], []
    
    def explorar_excel(self, ruta_archivo):
        """
        Proporciona información detallada sobre el contenido de un archivo Excel:
        nombres de hojas, tamaños, rangos de datos, etc.
        
        Args:
            ruta_archivo (str): Ruta al archivo Excel
        
        Returns:
            dict: Información detallada sobre el archivo Excel
        """
        info = {'hojas': {}}
        
        try:
            # Utilizamos openpyxl para obtener información detallada
            workbook = openpyxl.load_workbook(ruta_archivo, data_only=True, read_only=True)
            
            # Información general
            info['nombre_archivo'] = Path(ruta_archivo).name
            info['total_hojas'] = len(workbook.sheetnames)
            info['nombres_hojas'] = workbook.sheetnames
            
            # Información por hoja
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Determinar rango de datos utilizados (aproximado en modo read_only)
                min_row, min_col = 1, 1
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Calcular rango en notación de Excel
                min_col_letter = get_column_letter(min_col)
                max_col_letter = get_column_letter(max_col)
                rango_usado = f"{min_col_letter}{min_row}:{max_col_letter}{max_row}"
                
                # Guardar información de la hoja
                info['hojas'][sheet_name] = {
                    'dimensiones': f"{max_row} filas x {max_col} columnas",
                    'rango_usado': rango_usado,
                    'primera_fila_ejemplo': []
                }
                
                # Intentar leer la primera fila como ejemplo
                try:
                    # Leer solo la primera fila para mostrar como ejemplo
                    primera_fila = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                    info['hojas'][sheet_name]['primera_fila_ejemplo'] = primera_fila
                except:
                    # Si falla, indicar que no se pudo obtener
                    info['hojas'][sheet_name]['primera_fila_ejemplo'] = ["No disponible en modo read_only"]
            
            return info
            
        except Exception as e:
            raise ValueError(f"Error al explorar archivo Excel: {str(e)}")
    
    def leer_json(self, ruta_archivo, schema=None):
        import json, os
        try:
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                # Si el archivo es mayor a 1MB, usar procesamiento en streaming
                if os.path.getsize(ruta_archivo) > 10**6:
                    import ijson
                    # Se asume que la raíz es una lista de elementos, ajusta según necesidad
                    data = list(ijson.items(f, 'item'))
                else:
                    data = json.load(f)
            # Validación contra el esquema JSON si se proporciona
            if schema:
                from jsonschema import validate, ValidationError
                try:
                    validate(instance=data, schema=schema)
                except ValidationError as e:
                    raise ValueError("Error en validación de esquema JSON: {}".format(e))
            # Se retorna data, pudiendo ser una estructura anidada o lista
            return data
        except Exception as e:
            print("Error al leer el JSON:", e)
            return None

    def cargar_esquema_json(self, ruta_esquema=None, esquema_dict=None):
        """
        Carga un esquema JSON desde un archivo o diccionario.
        
        Args:
            ruta_esquema (str, optional): Ruta al archivo de esquema JSON
            esquema_dict (dict, optional): Esquema JSON como diccionario Python
            
        Returns:
            dict: El esquema JSON cargado
            
        Raises:
            ValueError: Si no se proporciona ni ruta ni diccionario,
                       o si hay un error al cargar el esquema
        """
        if ruta_esquema is None and esquema_dict is None:
            raise ValueError("Debe proporcionar una ruta de esquema o un esquema como diccionario")
        
        try:
            if esquema_dict is not None:
                # Validar que el esquema proporcionado sea válido
                jsonschema.Draft7Validator.check_schema(esquema_dict)
                return esquema_dict
            
            # Cargar desde archivo
            with open(ruta_esquema, 'r') as f:
                schema = json.load(f)
                
            # Validar que el esquema sea válido
            jsonschema.Draft7Validator.check_schema(schema)
            return schema
            
        except json.JSONDecodeError as e:
            raise ValueError(f"El archivo de esquema no contiene JSON válido: {str(e)}")
        except SchemaError as e:
            raise ValueError(f"El esquema JSON no es válido: {str(e)}")
        except Exception as e:
            raise ValueError(f"Error al cargar el esquema JSON: {str(e)}")
    
    def validar_datos_json(self, datos, schema, is_lines=False, options=None):
        """
        Valida datos JSON ya cargados contra un esquema.
        Útil para validar datos antes de escribirlos a un archivo.
        
        Args:
            datos: Datos JSON (ya parseados como dict o list)
            schema: Esquema para validación
            is_lines: Si es True, valida cada elemento por separado
            options: Opciones de validación
            
        Returns:
            bool: True si los datos son válidos
            
        Raises:
            ValidationError: Si los datos no cumplen con el esquema
            SchemaError: Si el esquema es inválido
        """
        if options is None:
            options = {'strict': True, 'format_check': True, 'additional_properties': True}
        
        try:
            self._validar_json_contra_esquema(datos, schema, is_lines, options)
            return True
        except ValidationError as e:
            if options['strict']:
                raise
            print(f"Advertencia: Datos no válidos según el esquema: {str(e)}")
            return False

    def _normalizar_dataframe_json_avanzado(self, df, config=None):
        """
        Normaliza estructuras JSON anidadas en un DataFrame de forma recursiva.
        
        Args:
            df (pandas.DataFrame): DataFrame a normalizar
            config (dict, optional): Configuración para la normalización:
                - max_depth (int): Profundidad máxima de normalización recursiva, por defecto 10
                - explode_arrays (bool): Si es True, expande arrays en múltiples filas, por defecto True
                - sep (str): Separador para nombres de columnas anidadas, por defecto '.'
                - meta_prefix (str): Prefijo para columnas de metadatos, por defecto 'meta_'
                - handle_errors (str): Cómo manejar errores ('ignore', 'warn', 'raise'), por defecto 'warn'
                - ignore_columns (list): Lista de columnas a ignorar
                - only_columns (list): Lista de columnas a procesar (exclusiva con ignore_columns)
                
        Returns:
            pandas.DataFrame: DataFrame con estructuras anidadas normalizadas
        
        Raises:
            ValueError: Si hay un error en la configuración o en el proceso de normalización
        """
        if config is None:
            config = {}
        
        # Valores predeterminados de configuración
        max_depth = config.get('max_depth', 10)
        explode_arrays = config.get('explode_arrays', True)
        separator = config.get('sep', '.')
        meta_prefix = config.get('meta_prefix', 'meta_')
        error_handling = config.get('handle_errors', 'warn')
        ignore_columns = config.get('ignore_columns', [])
        only_columns = config.get('only_columns', [])
        
        # Verificar configuración contradictoria
        if ignore_columns and only_columns:
            raise ValueError("No puede especificar tanto 'ignore_columns' como 'only_columns'")
        
        # Función recursiva interna para normalizar estructuras anidadas
        def normalizar_recursivo(df, nivel=0, parent_key=''):
            if nivel >= max_depth:
                return df
            
            # Copia para no modificar el original durante la iteración
            df_result = df.copy()
            
            # Columnas a procesar
            cols_to_process = df_result.columns
            if only_columns:
                cols_to_process = [c for c in cols_to_process if c in only_columns]
            if ignore_columns:
                cols_to_process = [c for c in cols_to_process if c not in ignore_columns]
            
            # Para registrar columnas a expandir (evitar modificar durante la iteración)
            columns_to_expand = {}
            
            # Primera pasada para identificar columnas a expandir
            for col in cols_to_process:
                # Saltar si ya tiene un prefijo especificado (para evitar múltiple procesamiento)
                if parent_key and col.startswith(f"{parent_key}{separator}"):
                    continue
                    
                # Verificar si la columna contiene objetos anidados o arrays
                if df_result[col].dtype == 'object':
                    # Muestra no nula para examinar
                    sample = df_result[col].dropna().iloc[0] if not df_result[col].dropna().empty else None
                    
                    if sample is not None:
                        if isinstance(sample, dict):
                            columns_to_expand[col] = 'dict'
                        elif isinstance(sample, list) and explode_arrays:
                            # Verificar si es lista de diccionarios o valores simples
                            if sample and all(isinstance(item, dict) for item in sample):
                                columns_to_expand[col] = 'list_dict'
                            else:
                                columns_to_expand[col] = 'list_simple'
            
            # Segunda pasada para expandir columnas identificadas
            for col, tipo in columns_to_expand.items():
                try:
                    if tipo == 'dict':
                        # Normalizar columnas de tipo diccionario
                        prefix = f"{col}{separator}" if not parent_key else f"{parent_key}{separator}{col}{separator}"
                        
                        # Extraer cada campo del diccionario como una nueva columna
                        json_expanded = pd.json_normalize(
                            df_result[col].dropna(),
                            sep=separator
                        )
                        
                        if not json_expanded.empty:
                            # Crear índice para unir correctamente los datos
                            json_expanded.index = df_result[col].dropna().index
                            
                            # Renombrar columnas con prefijo adecuado
                            json_expanded = json_expanded.add_prefix(prefix)
                            
                            # Unir con el dataframe original
                            df_result = pd.concat([df_result.drop(columns=[col]), json_expanded], axis=1)
                            
                    elif tipo == 'list_dict':
                        # Para listas de diccionarios, explotar la columna
                        prefix = f"{col}{separator}" if not parent_key else f"{parent_key}{separator}{col}{separator}"
                        
                        # Expandir la lista, creando una fila por cada elemento
                        # Necesitamos manejar índices para reconstruir correctamente
                        expanded_rows = []
                        
                        for idx, value in df_result[col].items():
                            if isinstance(value, list) and value:
                                # Crear un DataFrame para cada item en la lista
                                for item in value:
                                    if isinstance(item, dict):
                                        # Añadir una columna con el índice original
                                        item_df = pd.DataFrame([item])
                                        item_df['__original_index'] = idx
                                        expanded_rows.append(item_df)
                        
                        if expanded_rows:
                            # Combinar todos los DataFrames expandidos
                            all_expanded = pd.concat(expanded_rows, ignore_index=True)
                            
                            # Añadir metadatos útiles
                            all_expanded[f"{meta_prefix}source_column"] = col
                            
                            # Normalizar recursivamente esta estructura expandida
                            normalized_expanded = normalizar_recursivo(all_expanded, nivel + 1, prefix.rstrip(separator))
                            
                            # Crear una columna con los datos originales en formato JSON string
                            df_result[f"{meta_prefix}{col}_json"] = df_result[col].apply(
                                lambda x: json.dumps(x) if isinstance(x, list) else None
                            )
                            
                            # Eliminar la columna original que contenía la lista
                            df_result = df_result.drop(columns=[col])
                            
                            # Devolver resultado expandido en una estructura secundaria
                            # Esto evita mezclar filas con diferentes niveles de expansión
                            return {
                                'main': df_result,
                                'expanded': {col: normalized_expanded}
                            }
                    
                    elif tipo == 'list_simple':
                        # Para listas de valores simples, convertir a string JSON
                        df_result[col] = df_result[col].apply(
                            lambda x: json.dumps(x) if isinstance(x, list) else x
                        )
                
                except Exception as e:
                    if error_handling == 'raise':
                        raise ValueError(f"Error al normalizar columna {col}: {str(e)}")
                    elif error_handling == 'warn':
                        print(f"Advertencia: Error al normalizar columna {col}: {str(e)}")
                    # Si es 'ignore', continuamos sin hacer nada
            
            return df_result
        
        # Iniciar normalización recursiva
        resultado = normalizar_recursivo(df)
        
        # Si el resultado es un diccionario con expansiones, procesar y retornar
        if isinstance(resultado, dict):
            # Aquí podríamos retornar múltiples DataFrames o consolidar de alguna forma
            # Por defecto retornamos el DataFrame principal
            return resultado['main']
        
        return resultado
